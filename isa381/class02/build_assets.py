"""Generate the class02 side assets: Colab notebook, Kahoot import sheet, game QR code.
Run from class02/:  python build_assets.py
"""
import json
import segno
from openpyxl import Workbook

# ---------- 1. Colab notebook ----------
def md(src):
    return {"cell_type": "markdown", "metadata": {}, "source": src}

def code(src):
    return {"cell_type": "code", "metadata": {}, "execution_count": None, "outputs": [], "source": src}

cells = [
    md("# ISA 381, Class 02: Meet the Machine That Runs Your Code\n\n"
       "Every term from today's class is something you can ask this machine about. "
       "Run the cells in order with **Shift + Enter**.\n\n"
       "Before you start: **File > Save a copy in Drive**, so you are working on your own copy.\n\n"
       "Nothing here can break anything. One cell is *supposed* to fail; that is the point of it."),
    md("## 1. Hardware: the CPU\n\nColab gave your notebook a virtual machine on a Google server. "
       "The lines starting with `!` are commands to that machine's operating system (Linux), not Python. "
       "This one asks the CPU to identify itself, then Python asks how many processor cores you were given."),
    code("!grep -m 1 'model name' /proc/cpuinfo\n\nimport os\nprint(\"Cores available to you:\", os.cpu_count())"),
    md("## 2. Hardware: main memory (RAM)\n\nHow much main memory does this machine have, and how much is in use right now? "
       "`free -h` prints it in human-readable units (G = gigabytes)."),
    code("!free -h"),
    md("## 3. Hardware: secondary storage\n\nAnd how big is the disk attached to this machine? "
       "Remember: this disk is *temporary* to the virtual machine. Your notebook is safe because it is saved in Google Drive, "
       "a different piece of secondary storage."),
    code("!df -h /"),
    md("## 4. Software: the operating system and the interpreter\n\n"
       "Two pieces of *system software* are running your code right now: the operating system, and the Python interpreter."),
    code("import platform, sys\n\nprint(\"Operating system:\", platform.platform())\nprint(\"Python interpreter:\", sys.version)"),
    md("## 5. Break the syntax on purpose\n\nThe next cell is missing the colon after the `if` line. Run it and read the "
       "interpreter's message carefully. Then add the colon and run it again."),
    code("monthly_sales = [1200, 950, 1150]\n\nif sum(monthly_sales) > 3000\n    print(\"Above target\")"),
    md("Notice that **nothing ran**, not even the first line. The interpreter checks the syntax of the cell before executing it; "
       "a syntax error stops translation. (A *logic* error, like the East vs West weighting from last class, is different: "
       "the code runs fine and quietly gives a misleading answer.)"),
    md("## 6. Watch main memory forget\n\nRun the next cell. The variable `stock_value` now exists in the machine's main memory."),
    code("stock_value = (50 - 20) * 3\nprint(\"stock_value is\", stock_value)"),
    md("Now restart the machine's Python session: **Runtime > Restart session** (confirm when asked). "
       "That clears main memory. Then run the next cell *without* re-running the one above."),
    code("print(stock_value)"),
    md("`NameError`: as far as Python is concerned, `stock_value` never existed. Main memory is **volatile**.\n\n"
       "But look at the notebook itself: every cell is still here, because the notebook *file* lives in secondary storage (Drive). "
       "Re-run the cell that defines `stock_value` and it comes back. Programs are stored in secondary storage and copied into "
       "main memory each time they run; that is the whole story of this class."),
    md("## 7. Optional: peek at the interpreter's steps\n\nPython's `dis` module shows the small steps the interpreter turns one "
       "statement into on its way to the CPU. Each line below is one tiny operation, just like the LOAD / ADD / STORE instructions "
       "on the teaching machine in class. (These are the interpreter's own instructions, one level above the CPU's machine language.)"),
    code("import dis\n\ndis.dis(\"avg = total / 3\")"),
    md("## 8. Optional: bits and bytes in Python\n\nA few built-in functions let you see the byte behind a character."),
    code("print(ord(\"H\"))              # the number ASCII assigns to H\nprint(format(72, \"08b\"))    # that number as 8 bits\nprint(chr(0b01001101))      # which letter is 01001101?\nprint(bin(1200))            # 1200 in binary"),
    md("## Done\n\nYou have now seen every Chapter 1 term on a real machine: CPU, main memory, secondary storage, operating system, "
       "interpreter, syntax error, bits and bytes. Keep this notebook; it runs anywhere Colab does."),
]

nb = {
    "cells": cells,
    "metadata": {
        "colab": {"name": "isa381_class02_inside_the_machine.ipynb", "provenance": []},
        "kernelspec": {"display_name": "Python 3", "name": "python3"},
        "language_info": {"name": "python"},
    },
    "nbformat": 4,
    "nbformat_minor": 0,
}
with open("isa381_class02_inside_the_machine.ipynb", "w", encoding="utf-8") as f:
    json.dump(nb, f, indent=1, ensure_ascii=False)

# ---------- 2. Kahoot import spreadsheet (Kahoot's template layout: headers on row 8, questions from row 9, column B) ----------
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"
ws["B2"] = "ISA 381, Class 02 recap: last class in four questions"
ws["B3"] = "Import this file with Kahoot's spreadsheet importer (Create > Import spreadsheet)."
headers = ["Question - max 120 characters", "Answer 1 - max 75 characters", "Answer 2 - max 75 characters",
           "Answer 3 - max 75 characters", "Answer 4 - max 75 characters",
           "Time limit (sec) – 5, 10, 20, 30, 60, 90, 120, or 240 secs", "Correct answer(s) - choose at least one"]
for j, h in enumerate(headers):
    ws.cell(row=8, column=2 + j, value=h)
questions = [
    ("Which two words describe the kind of language Python is?",
     "High-level and interpreted", "Low-level and compiled", "High-level and compiled", "Low-level and interpreted", 20, 1),
    ("Last class, East vs West: why did the two 'average order value' answers disagree?",
     "One weighted months by number of orders", "Gemini used the wrong file", "East's data had a typo", "Python cannot compute averages", 20, 1),
    ("Google Colab is a hosted version of which tool?",
     "Jupyter Notebook", "Microsoft Excel", "Visual Studio", "GitHub", 20, 1),
    ("In the Aug 16 Job Scout harvest, which title group most often asked for Python?",
     "Analyst / analytics", "Data scientist", "Engineer / developer", "Other business titles", 20, 1),
]
for i, q in enumerate(questions):
    for j, v in enumerate(q):
        ws.cell(row=9 + i, column=2 + j, value=v)
wb.save("kahoot_class02_recap.xlsx")

# ---------- 3. QR code for the game ----------
URL = "https://fmegahed.github.io/isa381/class02/where_does_it_live.html"
segno.make(URL, error="m").save("game_qr.png", scale=8, border=2, dark="#1b1b1b", light="#ffffff")
print("wrote notebook, kahoot sheet, game_qr.png for", URL)
